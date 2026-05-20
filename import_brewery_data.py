#!/usr/bin/env python3
"""
Brewery Data Import Script
Imports real brewery data from 'Brewery Handover Sheets.xlsx' into the SQLite database.
Clears all dummy data first, preserving users, settings, and agent_rules.
"""

import os
import sys
import sqlite3
import shutil
from datetime import datetime, date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'brewery.db')
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Brewery Handover Sheets.xlsx')

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def clean_value(val):
    """Clean a cell value for database insertion"""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == '' or val.lower() in ('none', 'null', 'n/a', 'na', '-'):
            return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    if hasattr(val, 'strftime') and hasattr(val, 'hour'):
        # datetime.time object
        return val.strftime('%H:%M')
    return val


def clean_number(val, default=0):
    """Clean a numeric value"""
    val = clean_value(val)
    if val is None:
        return default
    try:
        if isinstance(val, str):
            val = val.replace(',', '').replace(' ', '')
            return float(val) if '.' in val else int(float(val))
        return float(val) if isinstance(val, float) else val
    except (ValueError, TypeError):
        return default


def clean_date(val):
    """Clean a date value to YYYY-MM-DD string"""
    val = clean_value(val)
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, str):
        # Handle various date formats
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(val.strip(), fmt).strftime('%Y-%m-%d')
            except ValueError:
                continue
        return val  # Return as-is if can't parse
    return str(val)


def read_sheet(wb, sheet_name, data_start_row=5):
    """Read a sheet and return list of dicts using row 2 as headers.
    Data starts at data_start_row (default 5, skipping description/name/types/required rows).
    """
    ws = wb[sheet_name]
    
    # Read headers from row 2
    headers = []
    for cell in ws[2]:
        h = str(cell.value).strip() if cell.value else ''
        # Remove type hints like "(YYYY-MM-DD)" from header for matching
        # but keep the base name
        headers.append(h)
    
    rows = []
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, values_only=True):
        # Skip empty rows
        if all(v is None or str(v).strip() == '' for v in row):
            continue
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = val
        rows.append(row_dict)
    
    return rows


def get_header_map(wb, sheet_name):
    """Get a map of header index to cleaned header name from row 2"""
    ws = wb[sheet_name]
    headers = {}
    for i, cell in enumerate(ws[2]):
        if cell.value:
            headers[i] = str(cell.value).strip()
    return headers


# ============================================================
# MAIN IMPORT
# ============================================================

def main():
    print("=" * 60)
    print("🍺 Brewery Data Import")
    print("=" * 60)
    
    # Check files exist
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Excel file not found: {EXCEL_PATH}")
        sys.exit(1)
    if not os.path.exists(DB_PATH):
        print(f"❌ Database not found: {DB_PATH}")
        sys.exit(1)
    
    # Backup database
    backup_path = f"{DB_PATH}.pre_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    shutil.copy2(DB_PATH, backup_path)
    print(f"✅ Database backed up to: {backup_path}")
    
    # Load Excel
    print(f"\n📖 Reading Excel file: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    print(f"   Found {len(wb.sheetnames)} sheets")
    
    # Connect to database
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = OFF")  # Disable for import
    cursor = conn.cursor()
    
    # Track IDs for foreign key resolution
    id_maps = {}  # {table: {name: id}}
    
    # ============================================================
    # PHASE 1: CLEAR DUMMY DATA
    # ============================================================
    print("\n🗑️  Clearing dummy data...")
    
    # Tables to clear (in dependency order - children first)
    tables_to_clear = [
        'batch_cogs', 'batch_costs', 'batch_ingredients',
        'order_items', 'order_lot_allocation',
        'lot_usage', 'lot_tracking', 'lots',
        'traceability_chain',
        'finished_goods_lots',
        'invoice_items', 'invoices',
        'quality_records',
        'sales_orders',
        'production_batches',
        'recipe_other_ingredients', 'recipe_mash_steps',
        'recipe_yeast', 'recipe_hops', 'recipe_fermentables',
        'recipes',
        'yeast_usage_log', 'yeast_viability_tests', 'yeast_propagations', 'yeast_inventory',
        'yeast_strains',
        'staff_schedule', 'training_records',
        'maintenance_schedule',
        'daily_tasks', 'briefing_log',
        'financial_transactions',
        'production_schedule',
        'notifications', 'user_notifications',
        'customers',
        'raw_materials',
        'products',
        'staff',
        'equipment',
        'packaging_materials',
        'agent_logs',
        'performance_metrics',
        'audit_log',
        'chat_history',
    ]
    
    # Preserved tables: users, settings, agent_rules, chat_sessions, user_preferences, user_permissions
    
    for table in tables_to_clear:
        try:
            cursor.execute(f"DELETE FROM {table}")
            print(f"   Cleared {table}")
        except sqlite3.OperationalError as e:
            print(f"   ⚠️  Skipped {table}: {e}")
    
    conn.commit()
    print("   ✅ Dummy data cleared")
    
    # ============================================================
    # PHASE 2: IMPORT MASTER DATA (No dependencies)
    # ============================================================
    print("\n📦 Phase 2: Importing master data...")
    
    # --- 2a. Products ---
    print("\n  📋 Products...")
    ws = wb['1. Products']
    products_map = {}
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        name = clean_value(row[0])
        if not name:
            continue
        style = clean_value(row[1])
        abv = clean_number(row[2], 0)
        ibu = clean_number(row[3], 0)
        description = clean_value(row[4])
        # B2B Keg price (column F=index 5)
        price_b2b_keg = clean_number(row[5], 0)
        # B2B Case price (column G=index 6)
        price_b2b_case = clean_number(row[6], 0)
        # Use B2B keg price as default price_per_unit
        price = price_b2b_keg if price_b2b_keg > 0 else price_b2b_case
        is_active = int(clean_number(row[9], 1)) if row[9] is not None else 1
        
        # Skip placeholder rows
        if name.strip().lower() in ('text', 'required', ''):
            continue
        
        cursor.execute("""
            INSERT INTO products (name, style, abv, ibu, description, price_per_unit, is_active)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (name, style, abv, ibu, description, price, is_active))
        
        pid = cursor.lastrowid
        products_map[name.strip().lower()] = pid
        # Also store by partial name for fuzzy matching
        products_map[name.split('(')[0].strip().lower()] = pid
        # Store first significant word for fuzzy matching
        words = name.strip().lower().split()
        if len(words) >= 2:
            products_map[' '.join(words[:2])] = pid
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} products")
    
    # --- 2b. Raw Materials ---
    print("\n  📋 Raw Materials...")
    ws = wb['2. Raw Materials']
    materials_map = {}
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        name = clean_value(row[0])
        if not name:
            continue
        category = clean_value(row[1])
        unit = clean_value(row[2])
        quantity = clean_number(row[3], 0)
        min_quantity = clean_number(row[4], 0)
        cost_per_unit = clean_number(row[5], 0)
        supplier = clean_value(row[6])
        origin = clean_value(row[7])
        expiry_date = clean_date(row[8])
        storage_location = clean_value(row[9])
        notes = clean_value(row[10])
        
        cursor.execute("""
            INSERT INTO raw_materials (name, category, unit, quantity, min_quantity,
                cost_per_unit, supplier, origin, expiry_date, storage_location, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, category, unit, quantity, min_quantity, cost_per_unit,
              supplier, origin, expiry_date, storage_location, notes))
        
        mid = cursor.lastrowid
        materials_map[name.strip().lower()] = mid
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} raw materials")
    
    # --- 2c. Staff ---
    print("\n  📋 Staff...")
    ws = wb['3. Staff']
    staff_map = {}
    count = 0
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        name = clean_value(row[0])
        if not name:
            continue
        position = clean_value(row[1])
        department = clean_value(row[2])
        phone = clean_value(row[3])
        email = clean_value(row[4])
        hire_date = clean_date(row[5])
        # Column 6 = Base salary
        salary = clean_number(row[6], 0)
        
        cursor.execute("""
            INSERT INTO staff (name, position, department, phone, email, hire_date, salary)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (name, position, department, phone, email, hire_date, salary))
        
        sid = cursor.lastrowid
        staff_map[name.strip().lower()] = sid
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} staff members")
    
    # --- 2d. Customers ---
    print("\n  📋 Customers...")
    ws = wb['4. Customers']
    customers_map = {}
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        # Column 0 = no (customer number), Column 1 = name
        name = clean_value(row[1])
        if not name:
            continue
        cust_type = clean_value(row[2])
        contact_person = clean_value(row[3])
        phone = clean_value(row[4])
        email = clean_value(row[5])
        address = clean_value(row[6])
        city = clean_value(row[7])
        # ward = row[8] - not in DB schema, skip
        tax_id = clean_value(row[9])
        credit_limit = clean_number(row[10], 0)
        payment_terms = clean_value(row[11])
        notes = clean_value(row[12])
        
        cursor.execute("""
            INSERT INTO customers (name, type, contact_person, phone, email, address,
                city, province, tax_id, credit_limit, payment_terms, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, cust_type, contact_person, phone, email, address,
              city, city, tax_id, credit_limit, payment_terms, notes))
        
        cid = cursor.lastrowid
        customers_map[name.strip().lower()] = cid
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} customers")
    
    # --- 2e. Equipment ---
    print("\n  📋 Equipment...")
    ws = wb['5. Equipment']
    equipment_map = {}
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        name = clean_value(row[0])
        if not name:
            continue
        equipment_type = clean_value(row[1])
        capacity = clean_number(row[2], 0)
        capacity_unit = clean_value(row[3])
        status = clean_value(row[4])
        last_cleaned = clean_date(row[5])
        notes = clean_value(row[6])
        
        cursor.execute("""
            INSERT INTO equipment (name, equipment_type, capacity, capacity_unit, status, last_cleaned, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (name, equipment_type, capacity, capacity_unit, status, last_cleaned, notes))
        
        eid = cursor.lastrowid
        equipment_map[name.strip().lower()] = eid
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} equipment items")
    
    # --- 2f. Yeast Strains ---
    print("\n  📋 Yeast Strains...")
    ws = wb['6. Yeast Strains']
    yeast_map = {}
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        name = clean_value(row[0])
        if not name:
            continue
        lab = clean_value(row[1])
        product_id = clean_value(row[2])
        yeast_type = clean_value(row[3])
        form = clean_value(row[4])
        att_min = clean_number(row[5], 0)
        att_max = clean_number(row[6], 0)
        flocculation = clean_value(row[7])
        min_temp = clean_number(row[8], 0)
        max_temp = clean_number(row[9], 0)
        alc_tolerance = clean_number(row[10], 0)
        description = clean_value(row[11])
        
        cursor.execute("""
            INSERT INTO yeast_strains (name, lab, product_id, yeast_type, form,
                attenuation_min, attenuation_max, flocculation, min_temp, max_temp,
                alcohol_tolerance, description)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, lab, product_id, yeast_type, form, att_min, att_max,
              flocculation, min_temp, max_temp, alc_tolerance, description))
        
        yid = cursor.lastrowid
        yeast_map[name.strip().lower()] = yid
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} yeast strains")
    
    # --- 2g. Packaging Materials ---
    print("\n  📋 Packaging Materials...")
    ws = wb['7. Packaging Materials']
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        name = clean_value(row[0])
        if not name:
            continue
        pkg_type = clean_value(row[1])
        quantity = clean_number(row[2], 0)
        unit = clean_value(row[3])
        cost_per_unit = clean_number(row[4], 0)
        supplier = clean_value(row[5])
        reorder_level = clean_number(row[6], 0)
        notes = clean_value(row[7])
        
        cursor.execute("""
            INSERT INTO packaging_materials (name, type, quantity, unit, cost_per_unit, supplier, reorder_level, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (name, pkg_type, quantity, unit, cost_per_unit, supplier, reorder_level, notes))
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} packaging materials")
    
    # ============================================================
    # PHASE 3: IMPORT RECIPES (Depends on products)
    # ============================================================
    print("\n📝 Phase 3: Importing recipes...")
    
    # --- 3a. Recipes ---
    print("\n  📋 Recipes...")
    ws = wb['8. Recipes']
    recipes_map = {}
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        product_name = clean_value(row[0])
        recipe_name = clean_value(row[1])
        if not recipe_name:
            continue
        style = clean_value(row[2])
        batch_size = clean_number(row[3], 2000)
        batch_size_unit = clean_value(row[4])
        boil_time = clean_number(row[5], 60)
        efficiency = clean_number(row[6], 75)
        og = clean_number(row[7], 0)
        fg = clean_number(row[8], 0)
        abv = clean_number(row[9], 0)
        ibu = clean_number(row[10], 0)
        srm = clean_number(row[11], 0)  # EBC column
        description = clean_value(row[12])
        notes = clean_value(row[13])
        
        # Resolve product_id
        product_id = None
        if product_name:
            product_id = products_map.get(product_name.strip().lower())
        
        cursor.execute("""
            INSERT INTO recipes (name, product_id, style, batch_size, batch_size_unit,
                boil_time, efficiency, og, fg, abv, ibu, srm, description, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (recipe_name, product_id, style, batch_size, batch_size_unit,
              boil_time, efficiency, og, fg, abv, ibu, srm, description, notes))
        
        rid = cursor.lastrowid
        recipes_map[recipe_name.strip().lower()] = rid
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} recipes")
    
    # --- 3b. Recipe Fermentables ---
    print("\n  📋 Recipe Fermentables...")
    ws = wb['9. Recipe Fermentables']
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        recipe_name = clean_value(row[0])
        material_name = clean_value(row[1])
        if not recipe_name or not material_name:
            continue
        
        recipe_id = recipes_map.get(recipe_name.strip().lower())
        material_id = materials_map.get(material_name.strip().lower())
        amount = clean_number(row[2], 0)
        unit = clean_value(row[3])
        percentage = clean_number(row[4], 0)
        potential = clean_number(row[5], 0)
        color = clean_number(row[6], 0)
        notes = clean_value(row[7])
        
        cursor.execute("""
            INSERT INTO recipe_fermentables (recipe_id, material_id, name, amount, unit,
                percentage, potential, color, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (recipe_id, material_id, material_name, amount, unit,
              percentage, potential, color, notes))
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} recipe fermentables")
    
    # --- 3c. Recipe Hops ---
    print("\n  📋 Recipe Hops...")
    ws = wb['10. Recipe Hops']
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        recipe_name = clean_value(row[0])
        hop_name = clean_value(row[1])
        if not recipe_name or not hop_name:
            continue
        
        recipe_id = recipes_map.get(recipe_name.strip().lower())
        material_id = materials_map.get(hop_name.strip().lower())
        amount = clean_number(row[2], 0)
        unit = clean_value(row[3])
        alpha_acid = clean_number(row[4], 0)
        boil_time = clean_number(row[5], 60)
        use_type = clean_value(row[6])
        ibu_contribution = clean_number(row[7], 0)
        notes = clean_value(row[8])
        
        cursor.execute("""
            INSERT INTO recipe_hops (recipe_id, material_id, name, amount, unit,
                alpha_acid, boil_time, use_type, ibu_contribution, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (recipe_id, material_id, hop_name, amount, unit,
              alpha_acid, boil_time, use_type, ibu_contribution, notes))
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} recipe hops")
    
    # --- 3d. Recipe Yeast ---
    print("\n  📋 Recipe Yeast...")
    ws = wb['11. Recipe Yeast']
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        recipe_name = clean_value(row[0])
        yeast_name = clean_value(row[1])
        if not recipe_name:
            continue
        
        recipe_id = recipes_map.get(recipe_name.strip().lower())
        yeast_id = yeast_map.get(str(yeast_name).strip().lower()) if yeast_name else None
        lab = clean_value(row[2])
        product_id_str = clean_value(row[3])
        form = clean_value(row[4])
        attenuation = clean_number(row[5], 75)
        min_temp = clean_number(row[6], 18)
        max_temp = clean_number(row[7], 22)
        notes = clean_value(row[8])
        
        cursor.execute("""
            INSERT INTO recipe_yeast (recipe_id, yeast_id, name, lab, product_id,
                form, attenuation, min_temp, max_temp, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (recipe_id, yeast_id, yeast_name or '', lab, product_id_str,
              form, attenuation, min_temp, max_temp, notes))
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} recipe yeast entries")
    
    # --- 3e. Recipe Mash Steps ---
    print("\n  📋 Recipe Mash Steps...")
    ws = wb['12. Recipe Mash Steps']
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        recipe_name = clean_value(row[0])
        if not recipe_name:
            continue
        
        recipe_id = recipes_map.get(recipe_name.strip().lower())
        step_number = int(clean_number(row[1], 1))
        name = clean_value(row[2])
        step_type = clean_value(row[3])
        temperature = clean_number(row[4], 0)
        duration = int(clean_number(row[5], 60))
        notes = clean_value(row[6])
        
        cursor.execute("""
            INSERT INTO recipe_mash_steps (recipe_id, step_number, name, step_type,
                temperature, duration, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (recipe_id, step_number, name, step_type, temperature, duration, notes))
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} recipe mash steps")
    
    # --- 3f. Recipe Other Ingredients ---
    print("\n  📋 Recipe Other Ingredients...")
    ws = wb['13. Recipe Other Ingredients']
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        recipe_name = clean_value(row[0])
        if not recipe_name:
            continue
        
        recipe_id = recipes_map.get(recipe_name.strip().lower())
        name = clean_value(row[1])
        ingredient_type = clean_value(row[2])
        amount = clean_number(row[3], 0)
        unit = clean_value(row[4])
        add_time = clean_value(row[5])
        notes = clean_value(row[6])
        
        cursor.execute("""
            INSERT INTO recipe_other_ingredients (recipe_id, name, ingredient_type,
                amount, unit, add_time, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (recipe_id, name, ingredient_type, amount, unit, add_time, notes))
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} recipe other ingredients")
    
    # ============================================================
    # PHASE 4: IMPORT PRODUCTION & SALES DATA
    # ============================================================
    print("\n🏭 Phase 4: Importing production & sales data...")
    
    # --- 4a. Production Batches ---
    print("\n  📋 Production Batches...")
    ws = wb['14. Production Batches']
    batches_map = {}  # batch_number -> id
    count = 0
    auto_created_products = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        batch_number = clean_value(row[0])
        product_name = clean_value(row[1])
        if not batch_number or not product_name:
            continue
        
        product_id = products_map.get(product_name.strip().lower())
        # Auto-create missing product if not found
        if not product_id:
            cursor.execute("""
                INSERT INTO products (name, style, abv, ibu, description, price_per_unit, is_active)
                VALUES (?, ?, 0, 0, ?, 0, 1)
            """, (product_name, product_name, f'Auto-created from batch data'))
            product_id = cursor.lastrowid
            products_map[product_name.strip().lower()] = product_id
            auto_created_products += 1
        
        tank_name = clean_value(row[2])
        tank_id = equipment_map.get(tank_name.strip().lower()) if tank_name else None
        planned_quantity = clean_number(row[3], 0)
        actual_quantity = clean_number(row[4], 0) if row[4] is not None else None
        status = clean_value(row[5]) or 'planned'
        start_date = clean_date(row[6])
        end_date = clean_date(row[7])
        brewer_name = clean_value(row[8])
        brewer_id = staff_map.get(brewer_name.strip().lower()) if brewer_name else None
        notes = clean_value(row[9])
        
        cursor.execute("""
            INSERT INTO production_batches (batch_number, product_id, tank_id,
                planned_quantity, actual_quantity, status, start_date, end_date,
                brewer_id, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (batch_number, product_id, tank_id, planned_quantity, actual_quantity,
              status, start_date, end_date, brewer_id, notes))
        
        bid = cursor.lastrowid
        batches_map[batch_number.strip().lower()] = bid
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} production batches")
    
    # --- 4b. Sales Orders ---
    print("\n  📋 Sales Orders...")
    ws = wb['15. Sales Orders']
    orders_map = {}  # order_number -> id
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        order_number = clean_value(row[0])
        customer_name = clean_value(row[1])
        if not order_number or not customer_name:
            continue
        
        customer_id = customers_map.get(customer_name.strip().lower())
        if not customer_id:
            # Try to find partial match
            for cname, cid in customers_map.items():
                if customer_name and cname in customer_name.strip().lower():
                    customer_id = cid
                    break
        if not customer_id:
            continue  # Skip if can't resolve customer
        
        order_date = clean_date(row[2])
        delivery_date = clean_date(row[3])
        status = clean_value(row[4]) or 'pending'
        payment_status = clean_value(row[5]) or 'unpaid'
        notes = clean_value(row[6])
        
        cursor.execute("""
            INSERT OR IGNORE INTO sales_orders (order_number, customer_id, order_date,
                delivery_date, status, payment_status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (order_number, customer_id, order_date, delivery_date,
              status, payment_status, notes))
        
        oid = cursor.lastrowid
        if oid:
            orders_map[order_number.strip().lower()] = oid
            count += 1
    conn.commit()
    print(f"    ✅ Imported {count} sales orders")
    
    # --- 4c. Order Items ---
    print("\n  📋 Order Items...")
    ws = wb['16. Order Items']
    count = 0
    skipped = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        order_number = clean_value(row[0])
        product_name = clean_value(row[1])
        if not order_number or not product_name:
            continue
        
        order_id = orders_map.get(order_number.strip().lower())
        if not order_id:
            skipped += 1
            continue
        
        # Try to resolve product by name (product names in orders may include packaging)
        product_id = None
        product_lower = product_name.strip().lower()
        # Try exact match first
        product_id = products_map.get(product_lower)
        if not product_id:
            # Try partial match - check if any product name is contained in the order product name
            for pname, pid in products_map.items():
                if pname in product_lower or product_lower in pname:
                    product_id = pid
                    break
            if not product_id:
                # Try matching first word
                first_word = product_lower.split('(')[0].strip().split()[0] if product_lower.split('(')[0].strip() else ''
                product_id = products_map.get(first_word)
        
        quantity = clean_number(row[2], 1)
        unit_price = clean_number(row[3], 0)
        discount = clean_number(row[4], 0)
        subtotal = quantity * unit_price * (1 - discount / 100) if discount else quantity * unit_price
        
        # Auto-create product if not found
        if not product_id:
            cursor.execute("""
                INSERT INTO products (name, style, abv, ibu, description, price_per_unit, is_active)
                VALUES (?, ?, 0, 0, ?, ?, 1)
            """, (product_name, product_name, f'Auto-created from order data', unit_price))
            product_id = cursor.lastrowid
            products_map[product_lower] = product_id
        
        cursor.execute("""
            INSERT INTO order_items (order_id, product_id, quantity, unit_price, discount, subtotal)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (order_id, product_id, quantity, unit_price, discount, subtotal))
        count += 1
    conn.commit()
    if skipped:
        print(f"    ⚠️  Skipped {skipped} items (order not found)")
    print(f"    ✅ Imported {count} order items")
    
    # --- 4d. Financial Transactions ---
    print("\n  📋 Financial Transactions...")
    ws = wb['17. Financial Transactions']
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        trans_date = clean_date(row[0])
        trans_type = clean_value(row[1])
        if not trans_type:
            continue  # Skip rows without type
        category = clean_value(row[2])
        amount = clean_number(row[3], 0)
        description = clean_value(row[4])
        payment_method = clean_value(row[5])
        reference_id = clean_value(row[6])
        reference_type = clean_value(row[7])
        
        cursor.execute("""
            INSERT INTO financial_transactions (transaction_date, type, category,
                amount, description, payment_method, reference_id, reference_type)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (trans_date, trans_type, category, amount, description,
              payment_method, reference_id, reference_type))
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} financial transactions")
    
    # --- 4e. Quality Records ---
    print("\n  📋 Quality Records...")
    ws = wb['18. Quality Records']
    count = 0
    skipped = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        batch_number = clean_value(row[0])
        check_type = clean_value(row[1])
        if not batch_number or not check_type:
            continue
        
        batch_id = batches_map.get(batch_number.strip().lower())
        if not batch_id:
            skipped += 1
            continue
        
        value = clean_value(row[2])
        unit = clean_value(row[3])
        passed = int(clean_number(row[4], 1))
        inspector = clean_value(row[5])
        notes = clean_value(row[6])
        
        cursor.execute("""
            INSERT INTO quality_records (batch_id, check_type, value, unit, passed, inspector, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (batch_id, check_type, value, unit, passed, inspector, notes))
        count += 1
    conn.commit()
    if skipped:
        print(f"    ⚠️  Skipped {skipped} records (batch not found)")
    print(f"    ✅ Imported {count} quality records")
    
    # ============================================================
    # PHASE 5: IMPORT SUPPORTING DATA
    # ============================================================
    print("\n📊 Phase 5: Importing supporting data...")
    
    # --- 5a. Yeast Inventory ---
    print("\n  📋 Yeast Inventory...")
    ws = wb['19. Yeast Inventory']
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        yeast_name = clean_value(row[0])
        if not yeast_name:
            continue
        
        yeast_id = yeast_map.get(yeast_name.strip().lower())
        lot_number = clean_value(row[1])
        quantity = clean_number(row[2], 0)
        unit = clean_value(row[3])
        viability = clean_number(row[4], 100)
        mfg_date = clean_date(row[5])
        expiry = clean_date(row[6])
        storage = clean_value(row[7])
        
        cursor.execute("""
            INSERT INTO yeast_inventory (yeast_id, lot_number, quantity, unit,
                viability, manufacture_date, expiry_date, storage_location)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (yeast_id, lot_number, quantity, unit, viability, mfg_date, expiry, storage))
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} yeast inventory entries")
    
    # --- 5b. Staff Schedule ---
    print("\n  📋 Staff Schedule...")
    ws = wb['20. Staff Schedule']
    count = 0
    skipped = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        staff_name = clean_value(row[0])
        schedule_date = clean_date(row[1])
        shift = clean_value(row[2])
        if not staff_name or not schedule_date or not shift:
            continue
        
        staff_id = staff_map.get(staff_name.strip().lower())
        if not staff_id:
            skipped += 1
            continue
        
        start_time = clean_value(row[3])
        end_time = clean_value(row[4])
        notes = clean_value(row[5])
        
        # Format times
        if start_time and isinstance(start_time, str) and ':' in start_time:
            start_time = start_time[:5]  # HH:MM
        if end_time and isinstance(end_time, str) and ':' in end_time:
            end_time = end_time[:5]
        
        cursor.execute("""
            INSERT INTO staff_schedule (staff_id, schedule_date, shift, start_time, end_time, notes)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (staff_id, schedule_date, shift, start_time, end_time, notes))
        count += 1
    conn.commit()
    if skipped:
        print(f"    ⚠️  Skipped {skipped} entries (staff not found)")
    print(f"    ✅ Imported {count} staff schedule entries")
    
    # --- 5c. Training Records ---
    print("\n  📋 Training Records...")
    ws = wb['21. Training Records']
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        staff_name = clean_value(row[0])
        training_date = clean_date(row[1])
        topic = clean_value(row[2])
        if not staff_name or not training_date or not topic:
            continue
        
        staff_id = staff_map.get(staff_name.strip().lower())
        if not staff_id:
            continue  # Skip if staff not found
        trainer = clean_value(row[3])
        duration_hours = clean_number(row[4], 0)
        result = clean_value(row[5])
        notes = clean_value(row[6])
        
        cursor.execute("""
            INSERT INTO training_records (staff_id, training_date, topic, trainer,
                duration_hours, result, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (staff_id, training_date, topic, trainer, duration_hours, result, notes))
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} training records")
    
    # --- 5d. Maintenance Schedule ---
    print("\n  📋 Maintenance Schedule...")
    ws = wb['22. Maintenance Checklist']
    count = 0
    skipped = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        equipment_name = clean_value(row[0])
        task_name = clean_value(row[1])
        task_type = clean_value(row[2])
        if not equipment_name or not task_name or not task_type:
            continue
        
        equipment_id = equipment_map.get(equipment_name.strip().lower())
        if not equipment_id:
            skipped += 1
            continue
        
        frequency_days = int(clean_number(row[3], 30))
        next_due = clean_date(row[4])
        assigned_to_name = clean_value(row[5])
        assigned_to = staff_map.get(assigned_to_name.strip().lower()) if assigned_to_name else None
        status = clean_value(row[6]) or 'scheduled'
        # row[7] = priority (not in DB schema, skip)
        # row[8] = estimated_duration_min (not in DB schema)
        last_completed = clean_date(row[9])
        # row[10] = completion_notes
        notes = clean_value(row[11])
        
        cursor.execute("""
            INSERT INTO maintenance_schedule (equipment_id, task_type, task_name,
                description, frequency_days, last_completed, next_due, assigned_to, status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (equipment_id, task_type, task_name, notes, frequency_days,
              last_completed, next_due, assigned_to, status, notes))
        count += 1
    conn.commit()
    if skipped:
        print(f"    ⚠️  Skipped {skipped} entries (equipment not found)")
    print(f"    ✅ Imported {count} maintenance schedule entries")
    
    # --- 5e. Asset List (New Table) ---
    print("\n  📋 Asset List...")
    # Create asset_list table if it doesn't exist
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS asset_list (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_name TEXT NOT NULL,
            asset_type TEXT,
            category TEXT,
            serial_number TEXT,
            manufacturer TEXT,
            model TEXT,
            purchase_date DATE,
            purchase_cost REAL DEFAULT 0,
            current_value REAL DEFAULT 0,
            condition TEXT DEFAULT 'good',
            location TEXT,
            assigned_to TEXT,
            warranty_expiry DATE,
            insurance_policy TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()
    
    ws = wb['23. Asset List']
    count = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        asset_name = clean_value(row[0])
        if not asset_name:
            continue
        asset_type = clean_value(row[1])
        category = clean_value(row[2])
        serial_number = clean_value(row[3])
        manufacturer = clean_value(row[4])
        model_name = clean_value(row[5])
        purchase_date = clean_date(row[6])
        purchase_cost = clean_number(row[7], 0)
        current_value = clean_number(row[8], 0)
        condition = clean_value(row[9])
        location = clean_value(row[10])
        assigned_to = clean_value(row[11])
        warranty_expiry = clean_date(row[12])
        insurance_policy = clean_value(row[13])
        notes = clean_value(row[14])
        
        cursor.execute("""
            INSERT INTO asset_list (asset_name, asset_type, category, serial_number,
                manufacturer, model, purchase_date, purchase_cost, current_value,
                condition, location, assigned_to, warranty_expiry, insurance_policy, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (asset_name, asset_type, category, serial_number, manufacturer,
              model_name, purchase_date, purchase_cost, current_value, condition,
              location, assigned_to, warranty_expiry, insurance_policy, notes))
        count += 1
    conn.commit()
    print(f"    ✅ Imported {count} assets")
    
    # ============================================================
    # PHASE 6: VERIFICATION
    # ============================================================
    print("\n" + "=" * 60)
    print("📊 Import Verification")
    print("=" * 60)
    
    verify_tables = [
        'products', 'raw_materials', 'staff', 'customers', 'equipment',
        'yeast_strains', 'packaging_materials', 'recipes',
        'recipe_fermentables', 'recipe_hops', 'recipe_yeast',
        'recipe_mash_steps', 'recipe_other_ingredients',
        'production_batches', 'sales_orders', 'order_items',
        'financial_transactions', 'quality_records',
        'yeast_inventory', 'staff_schedule', 'training_records',
        'maintenance_schedule', 'asset_list'
    ]
    
    total_rows = 0
    for table in verify_tables:
        try:
            cursor.execute(f"SELECT COUNT(*) FROM {table}")
            c = cursor.fetchone()[0]
            total_rows += c
            print(f"  ✅ {table}: {c} rows")
        except:
            print(f"  ⚠️  {table}: not found")
    
    # Check preserved tables
    cursor.execute("SELECT COUNT(*) FROM users")
    user_count = cursor.fetchone()[0]
    print(f"\n  🔒 Preserved - users: {user_count} rows")
    cursor.execute("SELECT COUNT(*) FROM settings")
    settings_count = cursor.fetchone()[0]
    print(f"  🔒 Preserved - settings: {settings_count} rows")
    
    print(f"\n  📊 Total imported rows: {total_rows}")
    
    # Re-enable foreign keys and optimize
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA optimize")
    
    conn.close()
    wb.close()
    
    print("\n" + "=" * 60)
    print("✅ Import complete!")
    print("=" * 60)


if __name__ == '__main__':
    main()
